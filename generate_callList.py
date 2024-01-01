import pandas as pd
#import calendar
# import numpy as np
# from datetime import timedelta
from openpyxl.utils import get_column_letter
# import tkinter as tk
# from tkinter import filedialog, messagebox
from datetime import datetime
#import os
from openpyxl.styles import Alignment, Font, Color, PatternFill
import random
#import io

### my own function imports
#from generator_gui import *
from import_excel import *
from call_interval_functions import *
from post_hoc_swap import *
from availability import *
#from weekends import *


random.seed(10) # Set seed 


def no_calls_next_to_fixed_calls(df):
    nrows, ncols = df.shape
    # Iterate over each row in the DataFrame
    for index, row in df.iterrows():
        # Enumerate the row to get cell index and value
        for i, cell in enumerate(row):
            # If the cell contains "1"
            if cell == "1":
                # Check the cell to the left, if it exists and is not "1" or "x"
                if i > 0 and df.iloc[index, i-1] not in ("1", "x"):
                    df.iloc[index, i-1] = "x_gen"
                # Check the cell two to the left, if it exists and is not "1" or "x"
                if i > 1 and df.iloc[index, i-2] not in ("1", "x"):
                    df.iloc[index, i-2] = "x_gen"
                # Check the cell to the right, if it exists and is not "1" or "x"
                if i < ncols - 1 and df.iloc[index, i+1] not in ("1", "x"):
                    df.iloc[index, i+1] = "x_gen"
                # Check the cell two to the right, if it exists and is not "1" or "x"
                if i < ncols - 2 and df.iloc[index, i+2] not in ("1", "x"):
                    df.iloc[index, i+2] = "x_gen"
    return df

def generateCallList(df, expected_total, weekend_priority, sat_priority, sun_priority, availDates, availWeekend, availSat, availSun, expected_total_original):
    # Initialize the callList with None values for each date
    callList = {date: None for date in df.columns}

    for date in df.columns:
        on_call_doc = df[df[date] == '1'].index.tolist()

        if on_call_doc:
            selected_doc = on_call_doc[0]
        else:
            available_docs = [i for i in range(len(df)) if df.at[i, date] not in ['x', 'x_gen']]
            if not available_docs:
                raise ValueError(f"No doctors available on {date.strftime('%Y-%m-%d')}")

            # Cache original_ratio computation and use a single loop
            ratios = {i: expected_total[i] / availDates[i] for i in available_docs}
            max_ratio_docs = [doc for doc, ratio in ratios.items() if ratio == max(ratios.values())]

            if len(max_ratio_docs) > 1:
                # Find the doctor(s) with the lowest expected total original value
                lowest_expected_total = min(expected_total_original[doc] for doc in max_ratio_docs)
                candidates = [doc for doc in max_ratio_docs if expected_total_original[doc] == lowest_expected_total]

                selected_doc = random.choice(candidates) if len(candidates) > 1 else candidates[0]
            else:
                selected_doc = max_ratio_docs[0]

        # Update the callList, DataFrame, and availabilities
        callList[date] = selected_doc
        update_after_assignment(df, selected_doc, expected_total, date, selected_doc in on_call_doc)
        decrement_availabilities(df, date, availDates, availWeekend, availSat, availSun)

    return callList


def map_call_list_to_df(df_original, callList, drName):
    # Create df_final with the same structure as df_original
    df_final = df_original.copy()

    # Replace "NA" with empty cells
    df_final = df_final.map(lambda x: '' if x == 'NA' else x)

    # Convert df_final column dates to string format for comparison
    df_final_dates = df_final.columns.astype(str)

    # Convert callList dates to the DataFrame's column date format
    # You'll need to know the incoming format of your callList dates to parse them properly
    # Let's assume they are in the format 'yyyy-mm-dd'
    formatted_callList = {pd.to_datetime(date).strftime('%Y-%m-%d'): doctor_index 
                          for date, doctor_index in callList.items()}

    # Map formatted_callList to df_final
    for date, doctor_index in formatted_callList.items():
        if pd.isnull(doctor_index) or doctor_index >= len(drName):
            continue

        # Check if the formatted date is in df_final's columns
        if date in df_final_dates:
            # Find the correct row for the doctor
            doctor_row = df_final.index[doctor_index]
            # Update the cell to indicate the doctor is on call
            df_final.at[doctor_row, date] = '1'

    return df_final


def add_headers_to_df_final(df_final, drName):

    print("till here 1")
    
    # Now create df_with_headers
    # Add headers: Assuming the first row of df_final is not part of the headers
    headers = ['Doctor'] + list(df_final.columns)
    df_with_headers = pd.DataFrame(columns=headers)
    
    # Add the 'Day of Week' header if the columns are datetimes
    try:
        day_of_week = ['Day of Week'] + [pd.to_datetime(date).strftime('%a') for date in df_final.columns]
    except ValueError:
        # Handle the case where the columns couldn't be parsed as dates
        day_of_week = ['Day of Week'] + ['' for _ in df_final.columns]
    
    df_with_headers.loc[len(df_with_headers)] = day_of_week
    
    # Add doctor names and rows from df_final to df_with_headers
    for i, (index, row) in enumerate(df_final.iterrows()):
        doctor_name = drName[i] if i < len(drName) else ''
        data_row = [doctor_name] + list(row)
        df_with_headers.loc[len(df_with_headers)] = data_row
    
    # Format the DataFrame columns as dates if they're not already
    df_with_headers.columns = ['Doctor'] + [pd.to_datetime(date).strftime('%Y-%m-%d') if not isinstance(date, str) else date for date in df_with_headers.columns[1:]]

    return df_with_headers


def mark_fixed_calls(df_original, df_final):
    # Ensure that df_final is a copy to not modify the original DataFrame passed
    df_final = df_final.copy()
    # Iterate through each cell in the DataFrame
    for row_index, row in df_original.iterrows():
        for col_name, value in row.items():
            # If the original DataFrame has a '1', mark 'F' in the final DataFrame
            if value == '1':
                df_final.at[row_index, col_name] = 'F'  # Using 'at' instead of 'iat'
    return df_final

def format_output_excel(writer, year, month, total_calls, total_sat_calls, total_sun_calls, three_day_calls, four_day_calls, five_day_calls, six_day_calls, seven_day_calls):
    start_column = 7  # e.g. Assuming 'D' column is the fourth column
    column_width = 3.5
    
    # Get the openpyxl worksheet object
    worksheet = writer.sheets['Sheet1']
    
    # Set the value for A1 and B1 cells to 'Year' and 'Month' respectively
    worksheet['A1'].value = 'Year:'
    worksheet['A2'].value = year
    worksheet['B1'].value = 'Month:'
    worksheet['B2'].value = month

    # Set the column width starting from the index specified by start_column
    for col_idx in range(start_column, worksheet.max_column + 1):
        col_letter = get_column_letter(col_idx)
        worksheet.column_dimensions[col_letter].width = column_width
        # Calculate the day of the month from the column index
        day_of_month = col_idx - start_column + 1
        # Check if the day is valid for the given year and month
        try:
            # If it's a valid date, write the day number to the third row (D3, E3, F3, ...)
            datetime(year, month, day_of_month)
            worksheet[f'{col_letter}3'].value = day_of_month
            worksheet[f'{col_letter}3'].alignment = Alignment(horizontal='left')
        except ValueError:
            # If the day is not valid (like February 30th), break out of the loop
            break

    # Set headers for AI, AJ, AK columns
    worksheet['AL4'].value = 'Total Calls'
    worksheet['AM4'].value = 'Total Saturday Calls'
    worksheet['AN4'].value = 'Total Sunday Calls'
    worksheet['AO4'].value = '3 day 1 call'
    worksheet['AP4'].value = '4 day 1 call'
    worksheet['AQ4'].value = '5 day 1 call'
    worksheet['AR4'].value = '6 day 1 call'
    worksheet['AS4'].value = '7 day 1 call'

    # Populate the AI, AJ, and AK columns with the corresponding data
    for idx, (total, sat, sun, three, four, five, six, seven) in enumerate(zip(total_calls, total_sat_calls, total_sun_calls, three_day_calls, four_day_calls, five_day_calls, six_day_calls, seven_day_calls), start=5):
        worksheet[f'AL{idx}'].value = total
        worksheet[f'AM{idx}'].value = sat
        worksheet[f'AN{idx}'].value = sun
        worksheet[f'AO{idx}'].value = three
        worksheet[f'AP{idx}'].value = four
        worksheet[f'AQ{idx}'].value = five
        worksheet[f'AR{idx}'].value = six
        worksheet[f'AS{idx}'].value = seven

    # Apply alignment and font to the entire sheet
    alignment = Alignment(horizontal='left')
    font = Font(name='Consolas')
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = alignment
            cell.font = font


def calculate_call_metrics(df_final, year, month):
    # Convert the column names to datetime
    df_final.columns = pd.to_datetime(df_final.columns)

    # # Filter the DataFrame for the given year and month
    # mask = (df_final.columns.year == year) & (df_final.columns.month == month)
    # df_month = df_final.loc[:, mask]

    # Initialize lists to store the call counts for each doctor
    total_calls = []
    total_sat_calls = []
    total_sun_calls = []
    three_day_calls = []
    four_day_calls = []
    five_day_calls = []
    six_day_calls = []
    seven_day_calls = []

    # Calculate the call intervals for all doctors using the original function
    call_intervals_matrix = calculate_call_intervals(df_final)

    # Iterate over each row (doctor) to calculate the metrics
    for index, row in df_final.iterrows():
        # Count the total number of calls (marked with "1")
        total_call_count = (row == "1").sum()
        total_calls.append(total_call_count)

        # Count the total number of calls on Saturdays
        saturdays = row.index[row.index.weekday == 5]
        total_sat_call_count = (row[saturdays] == "1").sum()
        total_sat_calls.append(total_sat_call_count)

        # Count the total number of calls on Sundays
        sundays = row.index[row.index.weekday == 6]
        total_sun_call_count = (row[sundays] == "1").sum()
        total_sun_calls.append(total_sun_call_count)

        # Extract individual day call counts from the intervals matrix for this doctor
        three_day_calls.append(call_intervals_matrix[index, 2])
        four_day_calls.append(call_intervals_matrix[index, 3])
        five_day_calls.append(call_intervals_matrix[index, 4])
        six_day_calls.append(call_intervals_matrix[index, 5])
        seven_day_calls.append(call_intervals_matrix[index, 6])

    return (total_calls, total_sat_calls, total_sun_calls, three_day_calls, four_day_calls,
            five_day_calls, six_day_calls, seven_day_calls)


# Function to select the file and execute the main program
def runFile(excel_file_path):
    # Execute the main program logic here with the selected path
    # ... [Your main program logic with the selected excel_file_path] ...

    # Call the function at program initialization
    df, drName, expected_total, weekend_priority, sat_priority, sun_priority, year, month = read_excel(excel_file_path)

    # Convert column names to datetime for easier manipulation ************* IMPORTANT!!! remember type name for columns now.
    df.columns = pd.to_datetime(df.columns)

    # Keep an original copy of df
    df_original = df.copy()

    # Assuming df is already defined and contains the call schedule
    availDates, availWeekend, availSat, availSun = calculate_availability(df)

    ### store the original expect_total
    expected_total_original = expected_total

    # set "x"'s before and after fixed ph calls to avoid more frequent than 3 day 1 call 
    df = no_calls_next_to_fixed_calls(df)

    #df, df_weekend, weekend_priority, availWeekend = prefill_weekend(df, year, month, weekend_priority, availWeekend)

    # 2nd time use this function: set "x"'s before and after WEEKENDS calls (assigned by program) to avoid more frequent than 3 day 1 call 
    #df = no_calls_next_to_fixed_calls(df)

    ### Minus all the pre-arranged calls from expected totals (and potentially randomly pre-arranged weekend calls on version 3)
    expected_total = initial_deduct_expected_totals(df, expected_total)

    # generate call list 
    callList = generateCallList(df, expected_total, weekend_priority, sat_priority, sun_priority, availDates, availWeekend, availSat, availSun, expected_total_original)

    print('Preliminary call list generated')

    df_final = map_call_list_to_df(df_original, callList, drName)

    # POST HOC SWAPS
    df_final = swap_calls_for_better_intervals(df_final, df_original)

    # Assuming df_final is your DataFrame with the correct structure
    # Example usage:
    total_calls, total_sat_calls, total_sun_calls, three_day_calls, four_days_calls, five_day_calls, six_day_calls, seven_day_calls = calculate_call_metrics(df_final, year, month)

    df_final_marked_fixed = mark_fixed_calls(df_original, df_final)
    df_with_headers = add_headers_to_df_final(df_final_marked_fixed, drName)

    expected_metrics = pd.DataFrame({           # for output purpose
        'expected_total': expected_total_original,
        # 'weekend_priority': weekend_priority,
        # 'sat_priority': sat_priority,
        # 'sun_priority': sun_priority
    })

    return df_with_headers, expected_metrics, year, month, total_calls, total_sat_calls, total_sun_calls, three_day_calls, four_days_calls, five_day_calls, six_day_calls, seven_day_calls 

    

# # Function to exit the application
# def quit_application():
    
#     root.quit()


############################ MAIN PROGRAM #################################

# # Now you have 'df', 'schedule', and all the arrays loaded and ready for use
# print(df)
# print(expected_total)
# print(weekend_priority)
# print(sat_priority)
# print(sun_priority)

# ##############

# if __name__ == "__main__":
#     # Main application window
#     root = tk.Tk()
#     root.title("UCH Call List Generator")

#     # Title label
#     title_label = tk.Label(root, text="UCH Call List Generator", font=("Helvetica", 16))
#     title_label.pack(side=tk.TOP, pady=10)

#     # Button to select the file
#     select_button = tk.Button(root, text="Select Excel File", command=runFile)
#     select_button.pack(side=tk.TOP, pady=10)

#     # Descriptive text label
#     description_label = tk.Label(root, text="Select the timetable template that you edited, to generate the actual call list.\n"
#                                             "It will put the call list excel into the same folder. It won't overwrite any existing files.\n\n"
#                                             "F stands for the fixed on-call dates.\n\n"
#                                             "Reasons for error: \n"
#                                             "    1. You didn't close the Excel file.\n"
#                                             "    2. You didn't fill in the expected_total for all doctors.\n"
#                                             "    3. You changed anything outside the expected_total column and the timetable cells. \n"
#                                             "       (you should create the template again using 01_create_template.exe if any change is needed.)\n")
#     description_label.pack(side=tk.TOP, pady=10)

#     # Button to quit the application
#     quit_button = tk.Button(root, text="Quit", command=quit_application)
#     quit_button.pack(side=tk.BOTTOM, pady=10)

#     # Footer label with author's name
#     footer_label = tk.Label(root, text="by Dr Wong Chung Wai William", font=("Helvetica", 10))
#     footer_label.pack(side=tk.BOTTOM, pady=5)

#     # Run the GUI event loop
#     root.mainloop()