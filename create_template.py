# import tkinter as tk
# from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font     #, Color, PatternFill
# import os
# from datetime import datetime
import calendar
from openpyxl.utils import get_column_letter

def create_timetable(year, month, doctors):
    # Ensure year and month are integers
    year = int(year)
    month = int(month)

    # Create an Excel workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Skip first two rows
    ws.append([])
    ws.append([])

    # Third and fourth row: Year and Month
    ws.append(["Year:", year, "", ""])
    ws.append(["Month:", month, "", ""])

    # A7 to D7: constant field names
    ws['A7'] = "expected_total"
    #ws['B7'] = "weekend_priority"
    #ws['C7'] = "sat_priority"
    #ws['D7'] = "sun_priority"

    # Use calendar to find the first day of the month and the number of days in the month
    first_weekday, num_days = calendar.monthrange(year, month)

    # Write the headers for the dates and days starting from column H
    ws['F6'] = "Date"
    ws['F7'] = "Day"

    # Add the 'num_Dr' label in cell A5
    ws['A5'] = "num_Dr"

    # Calculate and add the total number of doctors in cell B5
    total_doctors = len(doctors)
    ws['B5'] = total_doctors

    # Fill in the dates and days starting from column H
    for day in range(1, num_days + 1):
        # The column is G (column 7) plus the day of the month offset by -1
        col = get_column_letter(7 + (day - 1))
        ws[col + '6'] = day  # Set the date

        # Calculate the day of the week and set it
        day_of_week = calendar.day_name[calendar.weekday(year, month, day)]
        ws[col + '7'] = day_of_week[:3]  # Set the abbreviated day name

    # F column: Doctor's names
    for idx, name in enumerate(doctors, start=8):
        ws.cell(row=idx, column=6, value=name)

    # Apply alignment and font to the entire sheet
    alignment = Alignment(horizontal='left')
    font = Font(name='Consolas')
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = alignment
            cell.font = font
            
    # Set the font color for Saturday and Sunday
    for i in range(1, num_days + 1):
        col = get_column_letter(8 + (i - 1))
        if calendar.weekday(year, month, i) in [5, 6]:  # 5 and 6 correspond to Saturday and Sunday
            ws[col + '7'].font = Font(color="FF0000", name='Consolas')

    # Set the column width to 4 for columns starting from G onwards
    for col_idx in range(7, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 4

    return wb

# def create_button_clicked():
#     year = year_entry.get()
#     month = month_entry.get()
#     doctors = doctors_text.get("1.0", tk.END).strip().split("\n")
    
#     if not year.isdigit() or not month.isdigit():
#         messagebox.showerror("Error", "Year and month should be numbers.")
#         return

#     folder_path = filedialog.askdirectory()
#     if not folder_path:
#         messagebox.showerror("Error", "No folder selected.")
#         return

#     wb = create_timetable(year, month, doctors)

#     # Save the file
#     filename = f"template_{year}_{month}.xlsx"
#     filepath = os.path.join(folder_path, filename)
#     file_counter = 1

#     # Check if file exists and rename accordingly
#     while os.path.isfile(filepath):
#         filepath = os.path.join(folder_path, f"template_{year}_{month}-{file_counter}.xlsx")
#         file_counter += 1
    
#     wb.save(filepath)
#     messagebox.showinfo("Success", f"Timetable template created at {filepath}. You can quit this program after reading the instructions at the bottom.")


# def quit_button_clicked():
#     root.destroy()

# if __name__ == "__main__":
#     # Tkinter GUI setup
#     root = tk.Tk()
#     root.title("Timetable Template Generator")
#     tk.Label(root, text="This program creates a timetable template in Excel format.").pack()

#     tk.Label(root, text="Year:").pack()
#     year_entry = tk.Entry(root)
#     year_entry.pack()

#     tk.Label(root, text="Month:").pack()
#     month_entry = tk.Entry(root)
#     month_entry.pack()

#     tk.Label(root, text="Enter doctors' names (one per line):").pack()
#     doctors_text = tk.Text(root, height=10)
#     doctors_text.pack()

#     create_button = tk.Button(root, text="Create Timetable Template", command=create_button_clicked)
#     create_button.pack()

#     quit_button = tk.Button(root, text="Quit", command=quit_button_clicked)
#     quit_button.pack()


#     tk.Label(root, text="In the output Excel file:").pack()
#     tk.Label(root, text="- Enter 'x' for no-call requests.").pack()
#     tk.Label(root, text="- Enter '1' to assign an on-call shift to a particular doctor on a particular date.").pack() 
#     tk.Label(root, text="  Only enter '1' for at most one doctor per date.").pack()
#     tk.Label(root, text="- Fill in the expected total number of on-call shifts for each doctor. (under 'expected_total')").pack()
#     tk.Label(root, text="  These totals do not need to be exact :)").pack()
#     tk.Label(root, text="").pack() 
#     tk.Label(root, text="The program will try to meet these criteria as closely as possible when generating the schedule.").pack()
#     tk.Label(root, text="").pack()
#     tk.Label(root, text="After filling out the Excel file, run the second .exe file in the folder to generate the final on-call schedule.").pack()

#     tk.Label(root, text="").pack()
#     tk.Label(root, text="Created by Dr William Wong Chung Wai. UCH Radiology.").pack()

#     # Run the application
#     root.mainloop()